from flask import Flask, render_template, request, jsonify, send_from_directory, abort, g
import os
import re
import sqlite3
from datetime import datetime
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
import platform

app = Flask(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))


# =========================
# 경로 설정
# =========================

SERVER_DATA_DIR = "/home/opc/data/consent"
SERVER_UPLOAD_DIR = os.path.join(SERVER_DATA_DIR, "uploads", "consent_letters")

LOCAL_DATA_DIR = os.path.join(BASE_DIR, "data")
LOCAL_UPLOAD_DIR = os.path.join(BASE_DIR, "uploads", "consent_letters")

if platform.system() == "Windows":
    DATA_DIR = LOCAL_DATA_DIR
    UPLOAD_DIR = LOCAL_UPLOAD_DIR
else:
    DATA_DIR = SERVER_DATA_DIR
    UPLOAD_DIR = SERVER_UPLOAD_DIR

DB_PATH = os.path.join(DATA_DIR, "vessels.db")

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(UPLOAD_DIR, exist_ok=True)

ALLOWED_EXTENSIONS = {"pdf", "jpg", "jpeg", "png", "webp"}


# =========================
# DB 기본
# =========================
def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH, timeout=10)
        g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    db = sqlite3.connect(DB_PATH)
    db.execute("""
        CREATE TABLE IF NOT EXISTS vessels (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            management_company TEXT DEFAULT '',
            builder TEXT DEFAULT '',
            delivery_date TEXT DEFAULT '',
            next_dry_dock TEXT DEFAULT '',
            category TEXT DEFAULT 'AG 내',
            fujairah_consent TEXT DEFAULT '동의',
            yanbu_consent TEXT DEFAULT '동의',
            consent_letter TEXT DEFAULT '확보',
            voyage_plan TEXT DEFAULT '',
            ag_supply_plan TEXT DEFAULT '',
            crew_plan_status TEXT DEFAULT '불요',
            crew_count TEXT DEFAULT '',
            crew_date TEXT DEFAULT '',
            crew_port TEXT DEFAULT '',
            crew_plan_detail TEXT DEFAULT '',
            bonus_count TEXT DEFAULT '',
            bonus_amount TEXT DEFAULT '',
            latitude REAL,
            longitude REAL,
            consent_file TEXT DEFAULT ''
        )
    """)

    columns = [row[1] for row in db.execute("PRAGMA table_info(vessels)").fetchall()]

    if "management_company" not in columns:
        db.execute("ALTER TABLE vessels ADD COLUMN management_company TEXT DEFAULT ''")

    if "builder" not in columns:
        db.execute("ALTER TABLE vessels ADD COLUMN builder TEXT DEFAULT ''")

    if "delivery_date" not in columns:
        db.execute("ALTER TABLE vessels ADD COLUMN delivery_date TEXT DEFAULT ''")

    if "next_dry_dock" not in columns:
        db.execute("ALTER TABLE vessels ADD COLUMN next_dry_dock TEXT DEFAULT ''")

    if "category" not in columns:
        db.execute("ALTER TABLE vessels ADD COLUMN category TEXT DEFAULT 'AG 내'")

    if "ag_supply_plan" not in columns:
        db.execute("ALTER TABLE vessels ADD COLUMN ag_supply_plan TEXT DEFAULT ''")

    db.commit()
    db.close()


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


# =========================
# DB <-> JSON 변환
# =========================
def row_to_vessel_dict(row):
    return {
        "name": row["name"] or "",
        "managementCompany": row["management_company"] or "",
        "builder": row["builder"] or "",
        "deliveryDate": row["delivery_date"] or "",
        "nextDryDock": row["next_dry_dock"] or "",
        "category": row["category"] or "AG 내",
        "fujairahConsent": row["fujairah_consent"] or "",
        "yanbuConsent": row["yanbu_consent"] or "",
        "consentLetter": row["consent_letter"] or "",
        "voyagePlan": row["voyage_plan"] or "",
        "agSupplyPlan": row["ag_supply_plan"] or "",
        "crewPlanStatus": row["crew_plan_status"] or "",
        "crewCount": row["crew_count"] or "",
        "crewDate": row["crew_date"] or "",
        "crewPort": row["crew_port"] or "",
        "crewPlanDetail": row["crew_plan_detail"] or "",
        "bonusCount": row["bonus_count"] or "",
        "bonusAmount": row["bonus_amount"] or "",
        "latitude": row["latitude"],
        "longitude": row["longitude"],
        "consentFile": row["consent_file"] or "",
    }


def load_vessels():
    db = get_db()
    rows = db.execute("SELECT * FROM vessels ORDER BY name").fetchall()
    return [row_to_vessel_dict(row) for row in rows]


def get_vessel_by_name(name):
    db = get_db()
    row = db.execute(
        "SELECT * FROM vessels WHERE LOWER(name) = LOWER(?)",
        (name.strip(),)
    ).fetchone()
    return row


def normalize_vessel_data(data, old_vessel=None):
    old_vessel = old_vessel or {}

    return {
        "name": str(data.get("name", "")).strip(),
        "managementCompany": str(data.get("managementCompany", old_vessel.get("managementCompany", ""))).strip(),
        "builder": str(data.get("builder", old_vessel.get("builder", ""))).strip(),
        "deliveryDate": str(data.get("deliveryDate", old_vessel.get("deliveryDate", ""))).strip(),
        "nextDryDock": str(data.get("nextDryDock", old_vessel.get("nextDryDock", ""))).strip(),
        "category": str(data.get("category", old_vessel.get("category", "AG 내"))).strip() or "AG 내",
        "fujairahConsent": str(data.get("fujairahConsent", old_vessel.get("fujairahConsent", "동의"))).strip(),
        "yanbuConsent": str(data.get("yanbuConsent", old_vessel.get("yanbuConsent", "동의"))).strip(),
        "consentLetter": str(data.get("consentLetter", old_vessel.get("consentLetter", "확보"))).strip(),
        "voyagePlan": str(data.get("voyagePlan", old_vessel.get("voyagePlan", ""))).strip(),
        "agSupplyPlan": str(data.get("agSupplyPlan", old_vessel.get("agSupplyPlan", ""))).strip(),
        "crewPlanStatus": str(data.get("crewPlanStatus", old_vessel.get("crewPlanStatus", "불요"))).strip(),
        "crewCount": str(data.get("crewCount", old_vessel.get("crewCount", ""))).strip(),
        "crewDate": str(data.get("crewDate", old_vessel.get("crewDate", ""))).strip(),
        "crewPort": str(data.get("crewPort", old_vessel.get("crewPort", ""))).strip(),
        "crewPlanDetail": str(data.get("crewPlanDetail", old_vessel.get("crewPlanDetail", ""))).strip(),
        "bonusCount": str(data.get("bonusCount", old_vessel.get("bonusCount", ""))).strip(),
        "bonusAmount": str(data.get("bonusAmount", old_vessel.get("bonusAmount", ""))).strip(),
        "latitude": data.get("latitude", old_vessel.get("latitude", "")),
        "longitude": data.get("longitude", old_vessel.get("longitude", "")),
        "consentFile": old_vessel.get("consentFile", "")
    }


def upsert_vessel(vessel):
    db = get_db()
    db.execute("""
        INSERT INTO vessels (
            name, management_company, builder, delivery_date, next_dry_dock,
            category, fujairah_consent, yanbu_consent, consent_letter,
            voyage_plan, ag_supply_plan, crew_plan_status, crew_count, crew_date,
            crew_port, crew_plan_detail, bonus_count, bonus_amount,
            latitude, longitude, consent_file
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(name) DO UPDATE SET
            management_company = excluded.management_company,
            builder = excluded.builder,
            delivery_date = excluded.delivery_date,
            next_dry_dock = excluded.next_dry_dock,
            category = excluded.category,
            fujairah_consent = excluded.fujairah_consent,
            yanbu_consent = excluded.yanbu_consent,
            consent_letter = excluded.consent_letter,
            voyage_plan = excluded.voyage_plan,
            ag_supply_plan = excluded.ag_supply_plan,
            crew_plan_status = excluded.crew_plan_status,
            crew_count = excluded.crew_count,
            crew_date = excluded.crew_date,
            crew_port = excluded.crew_port,
            crew_plan_detail = excluded.crew_plan_detail,
            bonus_count = excluded.bonus_count,
            bonus_amount = excluded.bonus_amount,
            latitude = excluded.latitude,
            longitude = excluded.longitude,
            consent_file = excluded.consent_file
    """, (
        vessel["name"],
        vessel["managementCompany"],
        vessel["builder"],
        vessel["deliveryDate"],
        vessel["nextDryDock"],
        vessel["category"],
        vessel["fujairahConsent"],
        vessel["yanbuConsent"],
        vessel["consentLetter"],
        vessel["voyagePlan"],
        vessel["agSupplyPlan"],
        vessel["crewPlanStatus"],
        vessel["crewCount"],
        vessel["crewDate"],
        vessel["crewPort"],
        vessel["crewPlanDetail"],
        vessel["bonusCount"],
        vessel["bonusAmount"],
        vessel["latitude"],
        vessel["longitude"],
        vessel["consentFile"],
    ))
    db.commit()


# =========================
# 기타 유틸
# =========================
def get_asset_version():
    paths = [
        os.path.join(BASE_DIR, "templates", "index.html"),
        os.path.join(BASE_DIR, "templates", "report.html"),
        os.path.join(app.static_folder, "js", "app.js"),
        os.path.join(app.static_folder, "css", "style.css"),
        DB_PATH
    ]

    mtimes = []
    for path in paths:
        if os.path.exists(path):
            mtimes.append(str(int(os.path.getmtime(path))))

    return "-".join(mtimes) if mtimes else "1"


def normalize_report_value(value, default="-"):
    text = str(value or "").strip()
    return text if text else default


def filter_vessels_for_report(vessels, filter_name):
    filter_name = (filter_name or "all").strip()

    if filter_name == "ag":
        return [v for v in vessels if (v.get("category") or "AG 내") == "AG 내"]

    if filter_name == "bothArea":
        return [v for v in vessels if (v.get("category") or "AG 내") == "얀부, 푸자이라"]

    if filter_name == "other":
        return [v for v in vessels if (v.get("category") or "AG 내") == "그외 지역"]

    if filter_name == "fujairah":
        return [v for v in vessels if (v.get("fujairahConsent") or "").strip() == "동의"]

    if filter_name == "yanbu":
        return [v for v in vessels if (v.get("yanbuConsent") or "").strip() == "동의"]

    if filter_name == "crewConfirmed":
        return [v for v in vessels if (v.get("crewPlanStatus") or "").strip() == "확정"]

    if filter_name == "crewPending":
        return [v for v in vessels if (v.get("crewPlanStatus") or "").strip() == "미정"]

    return vessels


def build_report_rows(vessels):
    rows = []
    for vessel in vessels:
        rows.append({
            "name": normalize_report_value(vessel.get("name")),
            "managementCompany": normalize_report_value(vessel.get("managementCompany")),
            "builder": normalize_report_value(vessel.get("builder")),
            "deliveryDate": normalize_report_value(vessel.get("deliveryDate")),
            "nextDryDock": normalize_report_value(vessel.get("nextDryDock")),
            "category": normalize_report_value(vessel.get("category"), "AG 내"),
            "fujairahConsent": normalize_report_value(vessel.get("fujairahConsent")),
            "yanbuConsent": normalize_report_value(vessel.get("yanbuConsent")),
            "consentLetter": normalize_report_value(vessel.get("consentLetter")),
            "voyagePlan": normalize_report_value(vessel.get("voyagePlan")),
            "agSupplyPlan": normalize_report_value(vessel.get("agSupplyPlan")),
            "crewPlanStatus": normalize_report_value(vessel.get("crewPlanStatus"), "불요"),
            "crewCount": normalize_report_value(vessel.get("crewCount")),
            "crewDate": normalize_report_value(vessel.get("crewDate")),
            "crewPort": normalize_report_value(vessel.get("crewPort")),
            "crewPlanDetail": normalize_report_value(vessel.get("crewPlanDetail")),
        })
    return rows


def report_summary(vessels):
    def norm_text(value):
        return str(value or "").strip()

    def norm_category(value):
        v = norm_text(value)
        if v in ["AG 내", "얀부, 푸자이라", "그외 지역"]:
            return v
        return "AG 내"

    both_area_vessels = [
        v for v in vessels
        if norm_category(v.get("category")) == "얀부, 푸자이라"
    ]

    return {
        "total": len(vessels),
        "fujairah_yes": sum(
            1 for v in both_area_vessels
            if norm_text(v.get("fujairahConsent")) == "동의"
        ),
        "yanbu_yes": sum(
            1 for v in both_area_vessels
            if norm_text(v.get("yanbuConsent")) == "동의"
        ),
        "no_consent": sum(
            1 for v in vessels
            if norm_text(v.get("consentLetter")) in ["미확보", "진행중"]
        ),
        "crew_confirmed": sum(
            1 for v in vessels
            if norm_text(v.get("crewPlanStatus")) == "확정"
        ),
        "crew_pending": sum(
            1 for v in vessels
            if norm_text(v.get("crewPlanStatus")) == "미정"
        ),
    }


def normalize_name_for_match(name):
    return re.sub(r"\s+", " ", str(name or "").strip()).upper()


def excel_cell_str(value):
    if value is None:
        return ""
    return str(value).strip()


def safe_float(value):
    try:
        if value is None or str(value).strip() == "":
            return None
        return float(value)
    except Exception:
        return None


def normalize_degree_text(text):
    text = str(text or "").strip()
    text = text.replace("º", "°")
    text = text.replace("˚", "°")
    text = text.replace("’", "'")
    text = text.replace("‘", "'")
    text = text.replace("＇", "'")
    text = text.replace("“", '"')
    text = text.replace("”", '"')
    text = text.replace("″", '"')
    text = text.replace("／", "/")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def dms_to_decimal(direction, degrees, minutes=0.0, seconds=0.0):
    value = float(degrees) + float(minutes) / 60.0 + float(seconds) / 3600.0
    direction = str(direction or "").upper().strip()
    if direction in {"S", "W"}:
        value = -value
    return value


def parse_excel_datetime(value):
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        return value

    text = str(value).strip()
    if not text:
        return None

    patterns = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y",
    ]
    for fmt in patterns:
        try:
            return datetime.strptime(text, fmt)
        except Exception:
            continue
    return None


def parse_degree_minute_coordinate(degree, minute, hemisphere, coord_type):
    deg = safe_float(degree)
    minute_val = safe_float(minute)
    hemi = str(hemisphere or "").strip().upper()

    if deg is None or minute_val is None or hemi not in {"N", "S", "E", "W"}:
        return None

    value = abs(deg) + (minute_val / 60.0)

    if hemi in {"S", "W"}:
        value = -value

    if coord_type == "lat" and -90 <= value <= 90:
        return round(value, 6)
    if coord_type == "lon" and -180 <= value <= 180:
        return round(value, 6)

    return None


def normalize_header(text):
    return re.sub(r"[^a-z0-9가-힣]", "", str(text or "").strip().lower())


def find_header_index(headers, candidates):
    normalized = [normalize_header(h) for h in headers]
    candidate_set = {normalize_header(c) for c in candidates}
    for idx, header in enumerate(normalized):
        if header in candidate_set:
            return idx
    return None


def is_new_position_format(headers):
    def cell(idx):
        if idx >= len(headers):
            return ""
        return str(headers[idx] or "").strip().lower()

    return (
        cell(3) == "name"
        and cell(8) == "date(lt)"
        and cell(17) == "latitude"
        and cell(18) == "latitude"
        and cell(19) == "latitude"
        and cell(20) == "longitude"
        and cell(21) == "longitude"
        and cell(22) == "longitude"
    )


def parse_single_coord(token):
    if token is None:
        raise ValueError("좌표값이 없습니다.")

    text = normalize_degree_text(token).upper().replace(",", "")
    text = text.strip()

    if not text:
        raise ValueError("좌표값이 비어 있습니다.")

    m = re.match(r"^([NSEW])\s*([0-9]+(?:\.[0-9]+)?)\s*$", text)
    if m:
        return dms_to_decimal(m.group(1), m.group(2), 0, 0)

    m = re.match(r"^([NSEW])\s*([0-9]+(?:\.[0-9]+)?)°\s*([0-9]+(?:\.[0-9]+)?)'\s*$", text)
    if m:
        return dms_to_decimal(m.group(1), m.group(2), m.group(3), 0)

    m = re.match(r"^([NSEW])\s*([0-9]+(?:\.[0-9]+)?)°\s*([0-9]+(?:\.[0-9]+)?)'\s*([0-9]+(?:\.[0-9]+)?)\"\s*$", text)
    if m:
        return dms_to_decimal(m.group(1), m.group(2), m.group(3), m.group(4))

    m = re.match(r"^([0-9]+(?:\.[0-9]+)?)\s*([NSEW])\s*$", text)
    if m:
        return dms_to_decimal(m.group(2), m.group(1), 0, 0)

    m = re.match(r"^([+-]?[0-9]+(?:\.[0-9]+)?)\s*$", text)
    if m:
        return float(m.group(1))

    raise ValueError(f"좌표 해석 실패: {token}")


def parse_combined_position(position_text):
    text = normalize_degree_text(position_text).upper()

    if not text:
        raise ValueError("위치 문자열이 비어 있습니다.")

    if "/" in text:
        parts = [p.strip() for p in text.split("/") if p.strip()]
        if len(parts) != 2:
            raise ValueError("슬래시(/) 위치 형식이 올바르지 않습니다.")

        first = parse_single_coord(parts[0])
        second = parse_single_coord(parts[1])

        first_has_ns = bool(re.search(r"[NS]", parts[0]))
        first_has_ew = bool(re.search(r"[EW]", parts[0]))
        second_has_ns = bool(re.search(r"[NS]", parts[1]))
        second_has_ew = bool(re.search(r"[EW]", parts[1]))

        if first_has_ns:
            lat = first
            lon = second
        elif first_has_ew:
            lon = first
            lat = second
        elif second_has_ns:
            lon = first
            lat = second
        elif second_has_ew:
            lat = first
            lon = second
        else:
            if abs(first) <= 90 and abs(second) <= 180:
                lat = first
                lon = second
            else:
                lon = first
                lat = second

        return lat, lon

    pattern = r'([NSEW])\s*[0-9]+(?:\.[0-9]+)?(?:\s*°\s*[0-9]+(?:\.[0-9]+)?(?:\s*\'\s*[0-9]+(?:\.[0-9]+)?")?\'?)?'
    matches = list(re.finditer(pattern, text))

    if len(matches) >= 2:
        coord1 = matches[0].group(0).strip()
        coord2 = matches[1].group(0).strip()

        value1 = parse_single_coord(coord1)
        value2 = parse_single_coord(coord2)

        if coord1.startswith(("N", "S")):
            lat, lon = value1, value2
        else:
            lon, lat = value1, value2

        return lat, lon

    raise ValueError(f"지원되지 않는 위치 형식입니다: {position_text}")


def extract_position_from_row(row_dict):
    position_keys = ["위치", "좌표", "POSITION", "Position", "position", "LOCATION", "Location", "location"]
    lat_keys = ["위도", "LAT", "Lat", "lat", "LATITUDE", "Latitude", "latitude"]
    lon_keys = ["경도", "LON", "Lon", "lon", "LONGITUDE", "Longitude", "longitude"]

    for key in position_keys:
        if key in row_dict and excel_cell_str(row_dict.get(key)):
            return parse_combined_position(row_dict.get(key))

    lat_val = None
    lon_val = None

    for key in lat_keys:
        if key in row_dict and excel_cell_str(row_dict.get(key)):
            lat_val = parse_single_coord(row_dict.get(key))
            break

    for key in lon_keys:
        if key in row_dict and excel_cell_str(row_dict.get(key)):
            lon_val = parse_single_coord(row_dict.get(key))
            break

    if lat_val is not None and lon_val is not None:
        return lat_val, lon_val

    raise ValueError("위치 또는 위도/경도 컬럼을 찾지 못했습니다.")


def pick_latest_position_rows(sheet):
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {}, 0, 0, 0

    def parse_with_header_row(header_row_index):
        if len(rows) <= header_row_index:
            return {}, 0, 0, 0

        headers = [excel_cell_str(h) for h in rows[header_row_index]]
        data_rows = rows[header_row_index + 1:]

        if not headers:
            raise ValueError("헤더 행을 찾을 수 없습니다.")

        latest_by_name = {}
        total_rows = 0
        invalid_count = 0
        skipped_empty_name = 0

        if is_new_position_format(headers):
            for row in data_rows:
                if not row:
                    continue

                raw_name = row[3] if len(row) > 3 else None
                vessel_name = excel_cell_str(raw_name)

                if not vessel_name:
                    skipped_empty_name += 1
                    continue

                total_rows += 1
                dt_value = parse_excel_datetime(row[8] if len(row) > 8 else None)

                lat = parse_degree_minute_coordinate(
                    row[17] if len(row) > 17 else None,
                    row[18] if len(row) > 18 else None,
                    row[19] if len(row) > 19 else None,
                    "lat",
                )

                lon = parse_degree_minute_coordinate(
                    row[20] if len(row) > 20 else None,
                    row[21] if len(row) > 21 else None,
                    row[22] if len(row) > 22 else None,
                    "lon",
                )

                if lat is None or lon is None:
                    invalid_count += 1
                    continue

                key = normalize_name_for_match(vessel_name)
                item = {
                    "name": vessel_name,
                    "latitude": float(lat),
                    "longitude": float(lon),
                    "dt": dt_value,
                }

                current = latest_by_name.get(key)
                if current is None:
                    latest_by_name[key] = item
                    continue

                current_dt = current.get("dt")
                if dt_value and current_dt:
                    if dt_value >= current_dt:
                        latest_by_name[key] = item
                elif dt_value and not current_dt:
                    latest_by_name[key] = item
                elif not dt_value and not current_dt:
                    latest_by_name[key] = item

            return latest_by_name, total_rows, invalid_count, skipped_empty_name

        name_idx = find_header_index(headers, [
            "선명", "선박명", "ship name", "shipname", "vesselname", "vessel", "name"
        ])
        date_idx = find_header_index(headers, [
            "date", "일자", "날짜", "시간", "datetime", "updatedate", "updatetime"
        ])
        lat_idx = find_header_index(headers, [
            "latitude", "lat", "위도"
        ])
        lon_idx = find_header_index(headers, [
            "longitude", "lon", "lng", "경도"
        ])
        position_idx = find_header_index(headers, [
            "위치", "position", "pos", "location", "좌표"
        ])

        if name_idx is None:
            raise ValueError("엑셀 헤더에서 선명 또는 선박명을 찾을 수 없습니다.")

        if lat_idx is None and lon_idx is None and position_idx is None:
            raise ValueError("엑셀 헤더에서 위도/경도 또는 위치 컬럼을 찾을 수 없습니다.")

        for row in data_rows:
            if not row:
                continue

            row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            vessel_name = excel_cell_str(row[name_idx] if name_idx < len(row) else None)

            if not vessel_name:
                skipped_empty_name += 1
                continue

            total_rows += 1
            dt_value = None

            if date_idx is not None and date_idx < len(row):
                dt_value = parse_excel_datetime(row[date_idx])

            try:
                lat, lon = extract_position_from_row(row_dict)
            except Exception:
                invalid_count += 1
                continue

            if abs(lat) > 90 or abs(lon) > 180:
                invalid_count += 1
                continue

            key = normalize_name_for_match(vessel_name)
            item = {
                "name": vessel_name,
                "latitude": float(lat),
                "longitude": float(lon),
                "dt": dt_value,
            }

            current = latest_by_name.get(key)
            if current is None:
                latest_by_name[key] = item
                continue

            current_dt = current.get("dt")
            if dt_value and current_dt:
                if dt_value >= current_dt:
                    latest_by_name[key] = item
            elif dt_value and not current_dt:
                latest_by_name[key] = item
            elif not dt_value and not current_dt:
                latest_by_name[key] = item

        return latest_by_name, total_rows, invalid_count, skipped_empty_name

    try:
        result = parse_with_header_row(1)
        parsed_map, total_rows, invalid_count, skipped_empty_name = result
        if parsed_map or total_rows > 0:
            return result
    except ValueError:
        pass

    return parse_with_header_row(0)


# =========================
# 캐시 방지
# =========================
@app.after_request
def add_no_cache_headers(response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    response.headers["Surrogate-Control"] = "no-store"
    return response


# =========================
# 화면
# =========================
@app.route("/")
def index():
    return render_template("index.html", version=get_asset_version())


@app.route("/report")
def report():
    vessels = load_vessels()
    filter_name = request.args.get("filter", "all")
    filtered_vessels = filter_vessels_for_report(vessels, filter_name)

    summary = report_summary(filtered_vessels)
    rows = build_report_rows(filtered_vessels)

    filter_label_map = {
        "all": "전체",
        "ag": "AG 내",
        "bothArea": "얀부, 푸자이라",
        "other": "그외 지역",
        "fujairah": "푸자이라 동의",
        "yanbu": "얀부 동의",
        "crewConfirmed": "선원교대 확정",
        "crewPending": "선원교대 미정",
    }

    report_mode = "all"
    if filter_name == "ag":
        report_mode = "ag"
    elif filter_name == "bothArea":
        report_mode = "bothArea"
    elif filter_name == "other":
        report_mode = "other"

    return render_template(
        "report.html",
        rows=rows,
        summary=summary,
        generated_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        version=get_asset_version(),
        report_filter_label=filter_label_map.get(filter_name, "전체"),
        report_mode=report_mode
    )


# =========================
# API
# =========================
@app.route("/api/vessels", methods=["GET"])
def get_vessels_api():
    db = get_db()
    count = db.execute("SELECT COUNT(*) FROM vessels").fetchone()[0]
    vessels = load_vessels()
    return jsonify(vessels)


@app.route("/api/vessel", methods=["POST"])
def save_single_vessel():
    try:
        data = request.get_json(silent=True) or {}

        vessel_name = str(data.get("name", "")).strip()
        original_name = str(data.get("_originalName", "")).strip()

        if not vessel_name:
            return jsonify({"success": False, "message": "선박명이 없습니다."}), 400

        try:
            latitude = float(data.get("latitude"))
            longitude = float(data.get("longitude"))
        except (TypeError, ValueError):
            return jsonify({"success": False, "message": "위도 또는 경도 값이 올바르지 않습니다."}), 400

        old_row = None
        if original_name:
            row = get_vessel_by_name(original_name)
            if row:
                old_row = row_to_vessel_dict(row)

        if old_row is None:
            row = get_vessel_by_name(vessel_name)
            if row:
                old_row = row_to_vessel_dict(row)

        normalized = normalize_vessel_data(data, old_vessel=old_row)
        normalized["latitude"] = latitude
        normalized["longitude"] = longitude

        db = get_db()

        if original_name and original_name.lower() != vessel_name.lower():
            existing_original = db.execute(
                "SELECT * FROM vessels WHERE LOWER(name)=LOWER(?)",
                (original_name,)
            ).fetchone()

            if existing_original:
                db.execute("DELETE FROM vessels WHERE LOWER(name)=LOWER(?)", (original_name,))
                db.commit()

        upsert_vessel(normalized)

        return jsonify({"success": True, "message": "저장 완료"})
    except Exception as e:
        return jsonify({"success": False, "message": f"저장 중 오류: {str(e)}"}), 500


@app.route("/api/vessel/<path:vessel_name>", methods=["DELETE"])
def delete_single_vessel(vessel_name):
    try:
        db = get_db()
        cur = db.execute(
            "DELETE FROM vessels WHERE LOWER(name) = LOWER(?)",
            (vessel_name.strip(),)
        )
        db.commit()

        if cur.rowcount == 0:
            return jsonify({"success": False, "message": "삭제할 선박을 찾지 못했습니다."}), 404

        return jsonify({"success": True, "message": "삭제 완료"})
    except Exception as e:
        return jsonify({"success": False, "message": f"삭제 중 오류: {str(e)}"}), 500


@app.route("/api/upload-consent", methods=["POST"])
def upload_consent():
    try:
        vessel_name = request.form.get("vesselName", "").strip()
        file = request.files.get("file")

        if not vessel_name:
            return jsonify({"success": False, "message": "선박명이 없습니다."}), 400

        if not file or file.filename == "":
            return jsonify({"success": False, "message": "업로드할 파일이 없습니다."}), 400

        if not allowed_file(file.filename):
            return jsonify({"success": False, "message": "허용되지 않는 파일 형식입니다."}), 400

        row = get_vessel_by_name(vessel_name)
        if row is None:
            return jsonify({"success": False, "message": "해당 선박을 찾을 수 없습니다."}), 404

        ext = file.filename.rsplit(".", 1)[1].lower()
        safe_name = secure_filename(vessel_name.upper().replace(" ", "_"))
        new_filename = f"{safe_name}.{ext}"
        save_path = os.path.join(UPLOAD_DIR, new_filename)

        old_filename = str(row["consent_file"] or "").strip()
        if old_filename:
            old_path = os.path.join(UPLOAD_DIR, old_filename)
            if os.path.exists(old_path) and old_filename != new_filename:
                try:
                    os.remove(old_path)
                except Exception:
                    pass

        file.save(save_path)

        db = get_db()
        db.execute(
            "UPDATE vessels SET consent_file = ? WHERE LOWER(name) = LOWER(?)",
            (new_filename, vessel_name)
        )
        db.commit()

        return jsonify({
            "success": True,
            "message": "동의서 업로드 완료",
            "filename": new_filename
        })

    except Exception as e:
        return jsonify({"success": False, "message": f"업로드 중 오류: {str(e)}"}), 500


@app.route("/api/upload-positions", methods=["POST"])
def upload_positions():
    try:
        file = request.files.get("file")

        if not file or file.filename == "":
            return jsonify({"success": False, "message": "업로드할 엑셀 파일이 없습니다."}), 400

        if not file.filename.lower().endswith(".xlsx"):
            return jsonify({"success": False, "message": "xlsx 파일만 업로드할 수 있습니다."}), 400

        workbook = load_workbook(file, data_only=True)
        sheet = workbook.active

        latest_by_name, total_rows, invalid_count, skipped_empty_name = pick_latest_position_rows(sheet)

        db = get_db()
        rows = db.execute("SELECT name FROM vessels").fetchall()

        db_name_map = {normalize_name_for_match(r["name"]): r["name"] for r in rows}
        db_name_set = set(db_name_map.keys())

        updated_count = 0
        success_name_set = set()
        failed_name_set = set()

        for normalized_name, item in latest_by_name.items():
            if normalized_name not in db_name_set:
                continue

            real_name = db_name_map[normalized_name]
            lat = item["latitude"]
            lon = item["longitude"]

            if lat is None or lon is None or abs(lat) > 90 or abs(lon) > 180:
                failed_name_set.add(real_name)
                continue

            db.execute(
                "UPDATE vessels SET latitude = ?, longitude = ? WHERE name = ?",
                (float(lat), float(lon), real_name)
            )
            updated_count += 1
            success_name_set.add(real_name)

        db.commit()

        not_updated_vessels = sorted([
            real_name for real_name in db_name_map.values()
            if real_name not in success_name_set
        ])

        failed_vessels = sorted(list(failed_name_set))

        return jsonify({
            "success": True,
            "message": "위치 업데이트 완료",
            "updatedCount": updated_count,
            "failedVessels": failed_vessels,
            "notUpdatedVessels": not_updated_vessels
        })

    except Exception as e:
        return jsonify({"success": False, "message": f"위치 업데이트 중 오류: {str(e)}"}), 500


@app.route("/uploads/consent_letters/<path:filename>")
def uploaded_consent_file(filename):
    file_path = os.path.join(UPLOAD_DIR, filename)
    if not os.path.exists(file_path):
        abort(404)

    response = send_from_directory(UPLOAD_DIR, filename, as_attachment=False, conditional=False)
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


# =========================
# 시작
# =========================
init_db()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8000, debug=True)